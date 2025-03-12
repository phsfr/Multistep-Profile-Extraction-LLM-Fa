import os
import pickle
import re
from copy import copy

import openai
import pandas as pd
from hazm import SentenceTokenizer, WordTokenizer, Normalizer, Stemmer, Lemmatizer
from openpyxl import load_workbook
from owlready2 import get_ontology, destroy_entity
from persian_tools import digits

from multiStepMet_utils import get_nlu_prediction, get_slot_values
from pron_slot_value import find_gender, find_marital_stat, find_num_children, find_num_siblings, find_hobby, find_job, \
    find_residence, find_name_gender, find_age
from together import Together
from tenacity import (
    retry,
    stop_after_attempt,
    wait_random_exponential,
)

UNK = 'نامشخص'
ontology_path = "ontology/expo_perinfex_onto_inferred_V6.rdf"
ontology_ns = "http://www.semanticweb.org/asus/ontologies/2024/2/expo_perinfex_onto_inferred_V2#"
person_ins_name = "PersonInst"

perinfex_onto = get_ontology("file://" + ontology_path).load()
perinfex = perinfex_onto.get_namespace(ontology_ns)
if perinfex.Person(person_ins_name):
    destroy_entity(perinfex.Person(person_ins_name))
onto_person = perinfex.Person(person_ins_name)

sent_splitter = SentenceTokenizer()
word_tokenizer = WordTokenizer(join_verb_parts=False)
normalizer = Normalizer()
stemmer = Stemmer()
lemmatizer = Lemmatizer()

openai.api_key = "YOUR OPENAI API KEY"

os.environ["TOGETHER_API_KEY"] = "YOUR TOGETHER API KEY"


def send_together_req_to_find_contrast(prompt):
    client = Together(api_key=os.environ.get("TOGETHER_API_KEY"))
    response = client.chat.completions.create(
        model="meta-llama/Llama-3-70b-chat-hf",
        messages=[{"role": "user", "content": prompt}],
        temperature=1,
        max_tokens=1024,
        top_p=0.5,
        frequency_penalty=0,
        presence_penalty=0,
    )
    return response.choices[0].message.content


def find_contrast(input_sent_file, input_excel_file, model_type='llama', start_idx=0, stop_idx=1000):
    sentences = open(input_sent_file, 'r', encoding='utf-8').readlines()
    sheet_name = 'contrast_' + model_type
    workbook = load_workbook(input_excel_file)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        workbook.save(input_excel_file)
        predicted_data = []
    else:
        df_existing = pd.read_excel(input_excel_file, sheet_name=sheet_name)
        predicted_data = df_existing.values.tolist()
    for idx, sentence in enumerate(sentences):
        if idx < start_idx:
            continue
        if idx >= stop_idx:
            break
        sentence = sentence.strip('\n').strip()
        prompt = "In the following Persian sentence, identify any probable explicit or implicit semantic contrasts or contradictions. If there is no contrast or contradiction, state that none is present. Provide an explanation for any identified contrast or contradiction. Sentence: " + '"' + sentence + '"'
        if model_type == 'llama':
            response = send_together_req_to_find_contrast(prompt)
        elif 'gpt' in model_type:
            response = send_request_nlu(prompt, model_name='gpt-4o')
        # print(sentence)
        # print(response)
        # print('-' * 10)
        predicted_data.append((prompt, response))

    writer = pd.ExcelWriter(input_excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace')
    df = pd.DataFrame(predicted_data, columns=['prompt', 'response'])
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook.save(input_excel_file)
    writer.close()


def process_nlu_llm_response(response):
    slot_map = ['Sibl', 'Gend', 'Age', 'Job', 'Hobby', 'Child', 'Resist', 'Marri', 'Name']
    slot_values = {}
    qu_answers = response.strip().split('\n')
    utter_topics = []
    if len(qu_answers) < 10:
        return None, utter_topics
    for idx, answer in enumerate(qu_answers):
        if idx > 9:
            break
        # answer = answer.strip().replace('\u200C', "").replace(' ', "")
        if 'نمیدانم' in answer or 'نمی دانم' in answer or 'نمی‌دانم' in answer:
            continue
        answer = re.sub(r'^\d+[\.-]?\s*', '', answer)
        if idx == 0 or idx == 5:
            answer = answer.replace('#', '').replace('No', '').strip()
        if idx == 6:
            if answer.startswith('استان '):
                answer = answer.replace('استان ', '', 1).strip()
            elif answer.startswith('شهر '):
                answer = answer.replace('شهر ', '', 1).strip()
            elif answer.startswith('کشور '):
                answer = answer.replace('کشور ', '', 1).strip()

        if ',' in answer:
            separator = ','
        else:
            separator = '،'
        short_responses = answer.split(separator)
        if idx == 9:
            utter_topics = short_responses
        else:
            if idx == 0 or idx == 5:
                revised_short_responses = []
                for short_resp in short_responses:
                    short_resp = digits.convert_to_en(short_resp)
                    short_resp = re.sub(r'(\d)(\w+)', r'\1 \2', short_resp)
                    revised_short_responses.append(short_resp.strip())
                slot_values[slot_map[idx]] = revised_short_responses
            else:
                slot_values[slot_map[idx]] = short_responses

    return slot_values, utter_topics


def send_together_req(model, prompt, temperature=0, top_p=1.0):
    client = Together(api_key=os.environ.get("TOGETHER_API_KEY"))  # ))

    response = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=temperature,  # 1,
        max_tokens=1024,
        top_p=top_p,  # 0.5,
        frequency_penalty=0,
        presence_penalty=0,
    )
    return response.choices[0].message.content


@retry(wait=wait_random_exponential(min=1, max=60), stop=stop_after_attempt(6))
def send_request_nlu(prompt, model_name, top_p=1, temperature=1):
    response = openai.chat.completions.create(
        model=model_name,
        messages=[
            {"role": "user", "content": prompt}
        ],
        temperature=temperature,
        max_tokens=256,
        top_p=top_p,
        frequency_penalty=0,
        presence_penalty=0,
    )
    return response.choices[0].message.content


def send_nlu_llm_prompt(user_input, prev_utter, use_together=False):
    temperature = 0.0
    top_p = 1.0
    prompt = "Based on the user's input and the previous chatbot utterance in a Persian conversation, answer the following questions in the shortest possible form in Persian." + ' If the answer is not explicitly or nearly explicitly mentioned in the utterance, say "نمیدانم". '
    prompt += '\n' + """1. Has the user explicitly or nearly explicitly mentioned how many siblings they have? If yes, write in the form of "#No برادر", "#No خواهر", and "#No خواهربرادر". If not, say "نمیدانم"."""
    prompt += '\n' + """2. Has the user explicitly or nearly explicitly mentioned their gender? If yes, specify it just with 'مرد' or 'زن'. If not, say "نمیدانم". (do not imply gender from name)"""
    prompt += '\n' + """3. Has the user explicitly or nearly explicitly mentioned their age? If yes, what is it? (if birthdate with year was mentioned, just write year) If not, say "نمیدانم"."""
    prompt += '\n' + """4. Has the user explicitly or nearly explicitly mentioned their job? If yes, what is it? If not, say "نمیدانم". If there are multiple jobs mentioned, separate them with ','."""
    prompt += '\n' + """5. Has the user explicitly or nearly explicitly mentioned a hobby (an activity they can do in their leisure time)? If yes, what is it? If not, say "نمیدانم". If there are multiple hobbies mentioned, separate them with ','."""
    prompt += '\n' + """6. Has the user explicitly or nearly explicitly mentioned how many sons or daughters they have? write in the form of "#No پسر", "#No دختر", and "#No فرزند". If not, say "نمیدانم"."""
    prompt += '\n' + """7. Has the user explicitly or nearly explicitly mentioned their residence (district, city, province, or country)? If yes, what is it? If not, say "نمیدانم"."""
    prompt += '\n' + """8. Has the user explicitly or nearly explicitly mentioned if they are single or married? If yes, specify it just with 'متاهل' or 'مجرد'. If not, say "نمیدانم"."""
    prompt += '\n' + """9. Has the user explicitly or nearly explicitly mentioned their name? If yes, what is it? If not, say "نمیدانم". If there are multiple names mentioned, separate them with ','."""
    prompt += '\n' + """10. What is/are the main topic(s) in the user's utterance in order? If there are multiple main topics, separate them with ','. (The topic can only be one of these: "وضعیت تاهل", "اسم", "محل سکونت", "سرگرمی", "شغل", "سن", "جنسیت", "خانواده", "فرزندان", "خواهربرادران", "تحصیلات", "آب و هوا", "خداحافظی", "احوالپرسی". If it is outside of these, just answer "سایر".)"""
    prompt += '\n' + """**Note**: Distinguish clearly between the user's workplace and residence. Only consider information that is explicitly mentioned in the user's input and avoid making assumptions. If the information is not available, respond with "نمیدانم"."""
    prompt += '\n' + """**Note**: About question 1, if the user uses phrases like "من تک بچه ام", "من تک فرزندم", or "من تنها فرزند خانواده ام هستم", understand that they mean the user is an only child and does not have any siblings. These phrases do not indicate anything about the user's children."""
    prompt += '\n' + 'previous chatbot utterance: ' + prev_utter
    prompt += '\n' + "user's input: " + user_input

    if use_together:
        response = send_together_req("meta-llama/Meta-Llama-3.1-8B-Instruct-Turbo-128K", prompt, temperature, top_p)
    else:
        response = send_request_nlu(prompt, "gpt-4o", temperature, top_p)

    return prompt, response


def get_llm_nlu_analysis_result(inp_utter, dial_history, use_together=False, raw_response=None):
    if len(dial_history) == 0:
        prev_utter = 'شروع مکالمه'
    else:
        prev_utter = dial_history[-1].replace('chatbot: ', '').replace('user: ', '')
    if raw_response:
        response = raw_response
    else:
        prompt, response = send_nlu_llm_prompt(inp_utter, prev_utter, use_together)
    slot_values, nlu_topics = process_nlu_llm_response(response)
    if slot_values is None:
        print(f"Unable to process LLM NLU response: {response}")
        return UNK, '', response

    main_topic = ''
    for topic in nlu_topics:
        if topic == 'فرزندان' or topic == 'خواهربرادران':
            main_topic += 'خانواده' + ','
        else:
            main_topic += topic + ','
    main_topic = main_topic.strip(',')
    return slot_values, main_topic, response


def process_slot_values(inp_utter, slot_values, onto_ns):
    profile = {'topic': UNK, 'gender': UNK, 'marital status': UNK, 'numberOfGirls': UNK, 'minNumberOfGirls': UNK,
               'numberOfBoys': UNK, 'minNumberOfBoys': UNK, 'numberOfChild': UNK, 'minNumberOfChild': UNK,
               'minNumberOfSisters': UNK, 'numberOfSisters': UNK, 'minNumberOfBrothers': UNK, 'numberOfBrothers': UNK,
               'numberOfSibling': UNK, 'minNumberOfSibling': UNK, 'hobby': [], 'job': [], 'age': UNK, 'name': [],
               'city': UNK, 'country': UNK, 'region': UNK, 'province': UNK, 'section': UNK}
    for key, vals in slot_values.items():
        for val in vals:
            val = val.replace('.', '').replace('،', '').replace(',', '').replace('؛', '')
            if key == 'Gend':
                res = find_gender(val, stemmer, '')
                if res == 'Male': res = 'مرد'
                if res == 'Female': res = 'زن'
                profile['gender'] = res
            if key == 'Marri':
                res = find_marital_stat(val, lemmatizer, '')
                if res == 'Single': res = 'مجرد'
                if res == 'Married': res = 'متاهل'
                profile['marital status'] = res
            if key == 'Child':
                res = find_num_children(val, '')
                girlNum = res['girls']
                boyNum = res['boys']
                childNum = res['children']
                if girlNum != -1 and isinstance(girlNum, (int, str)):
                    prop = "numberOfGirls"
                    if isinstance(girlNum, str):
                        girlNum = int(girlNum.replace('+', ''))
                        prop = "minNumberOfGirls"
                    profile[prop] = girlNum

                if boyNum != -1 and isinstance(boyNum, (int, str)):
                    prop = "numberOfBoys"
                    if isinstance(boyNum, str):
                        boyNum = int(boyNum.replace('+', ''))
                        prop = "minNumberOfBoys"
                    profile[prop] = boyNum

                if childNum != -1 and isinstance(childNum, (int, str)):
                    prop = "numberOfChild"
                    if isinstance(childNum, str):
                        childNum = int(childNum.replace('+', ''))
                        prop = "minNumberOfChild"
                    profile[prop] = childNum

            if key == 'Sibl':
                res = find_num_siblings(val, '')
                sisterNum = res['sisters']
                bortherNum = res['brothers']
                siblNum = res['siblings']
                if sisterNum != -1 and isinstance(sisterNum, (int, str)):
                    prop = "numberOfSisters"
                    if isinstance(sisterNum, str):
                        sisterNum = int(sisterNum.replace('+', ''))
                        prop = "minNumberOfSisters"
                    profile[prop] = sisterNum
                if bortherNum != -1 and isinstance(bortherNum, (int, str)):
                    prop = "numberOfBrothers"
                    if isinstance(bortherNum, str):
                        bortherNum = int(bortherNum.replace('+', ''))
                        prop = "minNumberOfBrothers"
                    profile[prop] = bortherNum
                if siblNum != -1 and isinstance(siblNum, (int, str)):
                    prop = "numberOfSibling"
                    if isinstance(siblNum, str):
                        siblNum = int(siblNum.replace('+', ''))
                        prop = "minNumberOfSibling"
                    profile[prop] = siblNum

            if key == 'Hobby':
                hobby_value = find_hobby(val, onto_ns)
                profile['hobby'].append(hobby_value.label[0])

            if key == 'Job':
                res = find_job(val, onto_ns, lemmatizer)
                profile['job'].append(res.label[0])

            if key == 'Resist':
                res, val_type = find_residence(val, onto_ns, '')
                profile[val_type] = res[val_type.capitalize()]

            if key == 'Name':
                res = find_name_gender(val, inp_utter)
                profile['name'].append(res['Name'])

            if key == 'Age':
                res = find_age(val, inp_utter)
                if res != -1:
                    profile['age'] = res
    return profile


def get_processed_profile(inp_utter, dial_history, onto_ns, use_together=False, inp_response=None):
    slot_values, topic, raw_response = get_llm_nlu_analysis_result(inp_utter, dial_history, use_together, inp_response)
    profile = process_slot_values(inp_utter, slot_values, onto_ns)
    profile['topic'] = topic
    return profile, raw_response


def read_data(data_dir, output_file_name):
    map_tag_names = {'Marri': 'وضعیت تاهل', 'Name': 'اسم', 'Resist': 'محل سکونت',
                     'Hobby': 'سرگرمی', 'Sibl': 'تعداد خواهربرادر', 'Age': 'سن',
                     'Job': 'شغل', 'Child': 'تعداد فرزندان', 'Gend': 'جنسیت'}
    map_intent_names = {'job': 'شغل', 'hobby': 'سرگرمی', 'resistance': 'محل سکونت', 'family': 'خانواده',
                        'age': 'سن', 'gender': 'جنسیت', 'name': 'اسم', 'marriage': 'وضعیت تاهل', 'education': 'تحصیلات',
                        'greeting': 'احوالپرسی', 'goodby': 'خداحافظی', 'other': 'سایر', 'weather': 'آب و هوا'}

    main_dir = f'{data_dir}/test'
    label_lines = open(os.path.join(main_dir, 'label'), 'r', encoding='utf-8').readlines()
    sin_lines = open(os.path.join(main_dir, 'seq.in'), 'r', encoding='utf-8').readlines()
    sout_lines = open(os.path.join(main_dir, 'seq.out'), 'r', encoding='utf-8').readlines()
    gold_data = []
    for label, seq_in, seq_out in zip(label_lines, sin_lines, sout_lines):
        label = label.strip('\n').strip().split('\t')
        label_str = ''
        for la in label:
            label_str += map_intent_names[la] + ','
        label_str = label_str.strip(',')
        seq_out = seq_out.strip('\n').strip()
        seq_in = seq_in.strip('\n').strip()
        utterance = seq_in.split('\t')[0]
        prev_utter = seq_in.split('\t')[1]
        utter_slots = {'تعداد خواهربرادر': '-', 'جنسیت': '-', 'سن': '-', 'شغل': '-', 'سرگرمی': '-',
                       'تعداد فرزندان': '-', 'محل سکونت': '-', 'وضعیت تاهل': '-', 'اسم': '-'}
        tags = seq_out.split()

        if ['O'] * len(tags) == tags:
            gold_data.append((prev_utter, utterance, utter_slots, label_str))
        else:
            in_toks = utterance.split()
            assert len(in_toks) == len(tags)
            entity = ''
            for tok, tag in zip(in_toks, tags):
                if tag.startswith('B-'):
                    tag_name = tag[2:]
                    if len(entity) > 0:
                        if utter_slots[map_tag_names[tag_name]] == '-':
                            utter_slots[map_tag_names[tag_name]] = entity
                        else:
                            utter_slots[map_tag_names[tag_name]] += '|' + entity
                    entity = tok
                elif tag.startswith('I-'):
                    entity += ' ' + tok
                else:
                    if len(entity) > 0:
                        if utter_slots[map_tag_names[tag_name]] == '-':
                            utter_slots[map_tag_names[tag_name]] = entity
                        else:
                            utter_slots[map_tag_names[tag_name]] += '|' + entity
                    entity = ''
            if len(entity) > 0:
                if utter_slots[map_tag_names[tag_name]] == '-':
                    utter_slots[map_tag_names[tag_name]] = entity
                else:
                    utter_slots[map_tag_names[tag_name]] += '|' + entity
            gold_data.append((prev_utter, utterance, utter_slots, label_str))

    processed_gold = []
    for data in gold_data:
        slot_values = ''
        for key, val in data[2].items():
            slot_values += key + '=' + val + ','
        processed_gold.append((data[0], data[1], slot_values.strip(','), data[3]))
    df = pd.DataFrame(processed_gold, columns=['prev_utter', 'utterance', 'slot values', 'topic'])
    df.to_excel(output_file_name, index=False)

    with open('test_data.pkl', 'wb') as file:
        pickle.dump(processed_gold, file)


def send_together_req(model, prompt, temperature=0, top_p=1.0):
    client = Together(api_key=os.environ.get("TOGETHER_API_KEY"))  # ))

    response = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=temperature,  # 1,
        max_tokens=1024,
        top_p=top_p,
        frequency_penalty=0,
        presence_penalty=0,
    )
    return response.choices[0].message.content


def get_gold_preds(predict_file, predict_sheet, check_single_topic=False, type='mine_nlu'):
    df_existing = pd.read_excel(predict_file, sheet_name=predict_sheet)
    predicted_data = df_existing.values.tolist()
    fp, fn, tp = 0, 0, 0
    slot_metric = {'marital': {'tp': 0, 'fp': 0, 'fn': 0}, 'gender': {'tp': 0, 'fp': 0, 'fn': 0},
                   'siblings': {'tp': 0, 'fp': 0, 'fn': 0},
                   'age': {'tp': 0, 'fp': 0, 'fn': 0}, 'name': {'tp': 0, 'fp': 0, 'fn': 0},
                   'residence': {'tp': 0, 'fp': 0, 'fn': 0},
                   'job': {'tp': 0, 'fp': 0, 'fn': 0}, 'hobby': {'tp': 0, 'fp': 0, 'fn': 0},
                   'children': {'tp': 0, 'fp': 0, 'fn': 0}}
    slot_sklearn_metric = {'marital': {'tp': 0, 'fp': 0, 'fn': 0}, 'gender': {'tp': 0, 'fp': 0, 'fn': 0},
                           'siblings': {'tp': 0, 'fp': 0, 'fn': 0},
                           'age': {'tp': 0, 'fp': 0, 'fn': 0}, 'name': {'tp': 0, 'fp': 0, 'fn': 0},
                           'residence': {'tp': 0, 'fp': 0, 'fn': 0},
                           'job': {'tp': 0, 'fp': 0, 'fn': 0}, 'hobby': {'tp': 0, 'fp': 0, 'fn': 0},
                           'children': {'tp': 0, 'fp': 0, 'fn': 0}}
    fp_top, fn_top, tp_top = 0, 0, 0

    slot_map = {'جنسیت': 'gender', 'وضعیت تاهل': 'marital', 'سرگرمی': 'hobby', 'شغل': 'job', 'سن': 'age', 'اسم': 'name',
                'محل سکونت': 'residence', 'تعداد خواهربرادر': 'siblings', 'تعداد فرزندان': 'children'}
    slot_idx_map = ['siblings', 'gender', 'age', 'job', 'hobby', 'children', 'residence', 'marital', 'name']

    elem_slot = ['جنسیت =', 'سن =', 'شغل =', 'سرگرمی =', 'تعداد فرزندان =', 'محل سکونت =', 'وضعیت تاهل =', 'اسم =']
    y_true = set()
    y_pred = set()
    counter = 300
    for itr_idx, data in enumerate(predicted_data):
        if itr_idx == 40:
            same = True
        utterance = data[1]
        prev_utter = data[0]
        gold_tag = data[2]
        golden_tag = copy(gold_tag)
        gold_topic = data[3].split(',')
        gold_slots = gold_tag.strip().strip('|').split('|')
        pred_slots = []
        golden_slots = []
        pred_slots_str = data[4]

        pred_topic = data[5].replace('خواهربرادران', 'خانواده').replace('فرزندان', 'خانواده').replace('نامشخص',
                                                                                                      'آب و هوا').split(
            ',')
        stripped_pred_topics = [pred_t.strip() for pred_t in pred_topic]

        if check_single_topic:
            for t in stripped_pred_topics:
                if t in set(gold_topic):
                    tp_top += 1
                    break
        else:
            if set(stripped_pred_topics) == set(gold_topic):
                tp_top += 1

        for elem in elem_slot:
            next_idx = pred_slots_str.index(elem)
            next_idx_g = golden_tag.index(elem)
            slot_val = pred_slots_str[0:next_idx].split('=')[1]
            slot_val_g = golden_tag[0:next_idx_g].split('=')[1]
            pred_slots_str = pred_slots_str[next_idx:]
            golden_tag = golden_tag[next_idx_g:]
            pred_slots.append(slot_val.strip().strip('|').strip())
            golden_slots.append(slot_val_g.strip().strip('|').strip())
        pred_slots.append(pred_slots_str.split('=')[1].strip().strip('|').strip())
        golden_slots.append(golden_tag.split('=')[1].strip().strip('|').strip())

        binary_compare1 = []
        binary_compare2 = []
        for ix, p in enumerate(pred_slots):
            if p != '-':
                all_p_vals = p.split(',')
                all_g_vals = golden_slots[ix].split(',')
                existed_counters_g = 0
                for p_v in all_p_vals:
                    y_pred.add((slot_idx_map[ix], counter, counter))
                    if p_v in all_g_vals:
                        slot_sklearn_metric[slot_idx_map[ix]]['tp'] += 1
                        existed_counters_g += 1
                        y_true.add((slot_idx_map[ix], counter, counter))
                    else:
                        slot_sklearn_metric[slot_idx_map[ix]]['fp'] += 1
                    counter += 1
                if all_g_vals != ['-']:
                    remained_g_vals = len(all_g_vals) - existed_counters_g
                    for i in range(0, remained_g_vals):
                        slot_sklearn_metric[slot_idx_map[ix]]['fn'] += 1
                        y_true.add((slot_idx_map[ix], counter, counter))
                        counter += 1
            if p == '-' and golden_slots[ix] != '-':
                all_g_vals = golden_slots[ix].split(',')
                for g_v in all_g_vals:
                    slot_sklearn_metric[slot_idx_map[ix]]['fn'] += 1
                    y_true.add((slot_idx_map[ix], counter, counter))
                    counter += 1
            if golden_slots[ix] == p:
                if p != '-':
                    binary_compare2.append(1)
                    binary_compare1.append(1)
            else:
                binary_compare1.append(0)
                binary_compare2.append(1)

        for g_idx, gold_slot in enumerate(gold_slots):
            gold_val = gold_slot.split('=')[1].strip()
            slot_tag = gold_slot.split('=')[0].strip()
            if gold_val == '-':
                continue
            pred_val = pred_slots[g_idx]
            if pred_val == '-':
                fn += 1
                slot_metric[slot_map[slot_tag]]['fn'] += 1

        for idx, slot in enumerate(pred_slots):
            value = slot
            if value == '-':
                continue
            gold_val = gold_slots[idx].split('=')[1].strip()
            slot_tag = gold_slots[idx].split('=')[0].strip()
            if gold_val == '-':
                fp += 1
                slot_metric[slot_map[slot_tag]]['fp'] += 1
                continue
            if ',' in gold_val:
                gold_val = gold_val.split(',')
                pred_val = value.split(',')
                if set(gold_val) == set(pred_val):
                    tp += 1
                    continue
            if gold_val == value:
                tp += 1
                slot_metric[slot_map[slot_tag]]['tp'] += 1

    true_entities = y_true
    pred_entities = y_pred

    nb_correct = len(true_entities & pred_entities)
    nb_true = len(true_entities)

    recall_score = nb_correct / nb_true if nb_true > 0 else 0

    nb_pred = len(pred_entities)

    precision_score = nb_correct / nb_pred if nb_pred > 0 else 0

    p = nb_correct / nb_pred if nb_pred > 0 else 0
    r = nb_correct / nb_true if nb_true > 0 else 0
    f1_score = 2 * p * r / (p + r) if p + r > 0 else 0

    print(f'sklearn metric: precision = {precision_score}')
    print(f'sklearn metric: recall = {recall_score}')
    print(f'sklearn metric: F1 = {f1_score}')

    f1_avg, prec_avg, rec_agv = 0, 0, 0
    for slt in slot_sklearn_metric.keys():
        prec = slot_sklearn_metric[slt]['tp'] / (slot_sklearn_metric[slt]['tp'] + slot_sklearn_metric[slt]['fp'])
        rec = slot_sklearn_metric[slt]['tp'] / (slot_sklearn_metric[slt]['tp'] + slot_sklearn_metric[slt]['fn'])
        fone = 2 * (prec * rec) / (prec + rec)
        print(
            f"tp: {slot_sklearn_metric[slt]['tp']}  fp: {slot_sklearn_metric[slt]['fp']}   fn: {slot_sklearn_metric[slt]['fn']}")
        print(f'{slt}: precision = {prec}   recall = {rec}  f1 = {fone}')
        f1_avg += fone
        prec_avg += prec
        rec_agv += rec
    n = len(list(slot_sklearn_metric.keys()))
    print(f'avg F1 = {f1_avg / n}')
    print(f'avg precision = {prec_avg / n}')
    print(f'avg recall = {rec_agv / n}')


def refine_slot_values(pred_slots):
    child_val = ''
    sibl_val = ''
    resist_val = ''
    refined_pred_slots = []
    for ix in range(0, 2):
        refined_pred_slots.append(pred_slots[ix])
    for ix in range(2, 8):
        v = pred_slots[ix]
        if v != '-':
            if ix == 3 or ix == 5 or ix == 7:
                if pred_slots[ix - 1] == '-':
                    child_val += str(v) + ','
            else:
                child_val += str(v) + ','
    child_val = child_val.strip(',')
    if len(child_val) == 0:
        child_val = '-'
    refined_pred_slots.append(child_val)
    for ix in range(8, 14):
        v = pred_slots[ix]
        if v != '-':
            if ix == 8 or ix == 10:
                if pred_slots[ix + 1] == '-':
                    sibl_val += str(v) + ','
            elif ix == 13 and pred_slots[12] == '-':
                sibl_val += str(v) + ','
            else:
                sibl_val += str(v) + ','
    sibl_val = sibl_val.strip(',')
    if len(sibl_val) == 0:
        sibl_val = '-'
    refined_pred_slots.append(sibl_val)
    for ix in range(14, 18):
        refined_pred_slots.append(pred_slots[ix])
    for ix in range(18, len(pred_slots)):
        if pred_slots[ix] != '-':
            resist_val += str(pred_slots[ix]) + ','
    resist_val = resist_val.strip(',')
    if len(resist_val) == 0:
        resist_val = '-'
    refined_pred_slots.append(resist_val)
    return refined_pred_slots


def get_gold_preds_llm(predict_file, predict_sheet, check_single_topic=False):
    df_existing = pd.read_excel(predict_file, sheet_name=predict_sheet)
    predicted_data = df_existing.values.tolist()
    slot_sklearn_metric = {'marital': {'tp': 0, 'fp': 0, 'fn': 0}, 'gender': {'tp': 0, 'fp': 0, 'fn': 0},
                           'siblings': {'tp': 0, 'fp': 0, 'fn': 0},
                           'age': {'tp': 0, 'fp': 0, 'fn': 0}, 'name': {'tp': 0, 'fp': 0, 'fn': 0},
                           'residence': {'tp': 0, 'fp': 0, 'fn': 0},
                           'job': {'tp': 0, 'fp': 0, 'fn': 0}, 'hobby': {'tp': 0, 'fp': 0, 'fn': 0},
                           'children': {'tp': 0, 'fp': 0, 'fn': 0}}
    fp_top, fn_top, tp_top = 0, 0, 0

    slot_idx_map = ['gender', 'marital', 'children', 'siblings', 'hobby', 'job', 'age', 'name', 'residence']

    elem_slot_llm = ["تاهل =", "#دختران =", "#حداقل دختران =", "#پسران =", "#حداقل پسران =", "#فرزندان =",
                     "#حداقل فرزندان =", "#حداقل خواهران =", "#خواهران =", "#حداقل برادران =", "#برادران =",
                     "#خواهربرادران =", "#حداقل خواهربرادران =", "تفریح =", "شغل =", "سن =", "نام =", "شهر =", "کشور =",
                     "جهت جغرافیا =", "استان =", "بخش ="]
    y_true = set()
    y_pred = set()
    counter = 300

    # gpt 4o hobby fp: 99, fn: 35 | job fp: 6, fn: 4 | name fp: 3, fn: 5 | siblings: fp: 2 | children fp: 1 | residence fp: 4| gender fp: 5
    corrections = {'gender': {'fp': 5, 'fn': 0}, 'marital': {'fp': 0, 'fn': 0}, 'children': {'fp': 1, 'fn': 0}
        , 'siblings': {'fp': 2, 'fn': 0}, 'hobby': {'fp': 99, 'fn': 35}, 'job': {'fp': 6, 'fn': 4}
        , 'age': {'fp': 0, 'fn': 0}, 'name': {'fp': 3, 'fn': 5}, 'residence': {'fp': 4, 'fn': 0}}
    # phi-4 hobby fp: 104, fn: 96 | job fp: 24, fn: 13 | name fp: 1, fn: 0 | siblings: fp: 3 | children fp: 3 | residence fp: 6, fn: 2| gender fp: 2
    # corrections = {'gender': {'fp': 2, 'fn': 0}, 'marital': {'fp': 0, 'fn': 0}, 'children': {'fp': 3, 'fn': 0}
    #     , 'siblings': {'fp': 3, 'fn': 0}, 'hobby': {'fp': 104, 'fn': 96}, 'job': {'fp': 24, 'fn': 13}
    #     , 'age': {'fp': 0, 'fn': 0}, 'name': {'fp': 1, 'fn': 0}, 'residence': {'fp': 6, 'fn': 2}}
    # llama-3.1-8B hobby fp: 29, fn: 8 | job fp: 3, fn: 2 | marital fp: 1, fn: 0 | siblings: fp: 0 | children fp: 0 | residence fp: 1, fn: 0| gender fp: 1
    # corrections = {'gender': {'fp': 1, 'fn': 0}, 'marital': {'fp': 1, 'fn': 0}, 'children': {'fp': 0, 'fn': 0}
    #                   , 'siblings': {'fp': 0, 'fn': 0}, 'hobby': {'fp': 29, 'fn': 8}, 'job': {'fp': 3, 'fn': 2}
    #                   , 'age': {'fp': 0, 'fn': 0}, 'name': {'fp': 0, 'fn': 0}, 'residence': {'fp': 1, 'fn': 0}}

    # corrections = {'gender': {'fp': 0, 'fn': 0}, 'marital': {'fp': 0, 'fn': 0}, 'children': {'fp': 0, 'fn': 0}
    #                   , 'siblings': {'fp': 0, 'fn': 0}, 'hobby': {'fp': 0, 'fn': 0}, 'job': {'fp': 0, 'fn': 0}
    #                   , 'age': {'fp': 0, 'fn': 0}, 'name': {'fp': 0, 'fn': 0}, 'residence': {'fp': 0, 'fn': 0}}
    # ----------------------------------

    for itr_idx, data in enumerate(predicted_data):
        utterance = data[1]
        prev_utter = data[0]
        gold_tag = data[2]
        golden_tag = copy(gold_tag)
        gold_topic = data[3].split(',')
        pred_slots = []
        golden_slots = []
        pred_slots_str = data[4]
        t_seperator = ','
        if predict_sheet == 'phi-4':
            t_seperator = '،'
            if ',' in data[5]:
                t_seperator = ','
        pred_topic = data[5].replace('خواهربرادران', 'خانواده').replace('فرزندان', 'خانواده').replace('نامشخص',
                                                                                                      'آب و هوا').split(
            t_seperator)
        stripped_pred_topics = [pred_t.strip() for pred_t in pred_topic]

        if check_single_topic:
            for t in stripped_pred_topics:
                if t in set(gold_topic):
                    tp_top += 1
                    break
        else:
            if set(stripped_pred_topics) == set(gold_topic):
                tp_top += 1

        for elem in elem_slot_llm:
            next_idx = pred_slots_str.index(elem)
            next_idx_g = golden_tag.index(elem)
            slot_tag_g = golden_tag[0:next_idx_g].split('=')[0].strip()
            assert slot_tag_g == slot_tag_g
            slot_val = pred_slots_str[0:next_idx].split('=')[1]
            slot_val_g = golden_tag[0:next_idx_g].split('=')[1]
            pred_slots_str = pred_slots_str[next_idx:]
            golden_tag = golden_tag[next_idx_g:]
            pred_slots.append(slot_val.strip().strip('|').strip())
            golden_slots.append(slot_val_g.strip().strip('|').strip())
        pred_slots.append(pred_slots_str.split('=')[1].strip().strip('|').strip())
        golden_slots.append(golden_tag.split('=')[1].strip().strip('|').strip())
        assert len(pred_slots) == len(golden_slots)

        refined_pred_slots = refine_slot_values(pred_slots)
        refined_gold_slots = refine_slot_values(golden_slots)

        for ix, p in enumerate(refined_pred_slots):
            #
            slot_name = slot_idx_map[ix]
            if p != '-':
                all_p_vals = [pr.strip() for pr in p.split(',')]
                all_g_vals = [r.strip() for r in refined_gold_slots[ix].split(',')]
                existed_counters_g = 0
                for p_v in all_p_vals:
                    y_pred.add((slot_idx_map[ix], counter, counter))
                    if p_v in all_g_vals:
                        slot_sklearn_metric[slot_idx_map[ix]]['tp'] += 1
                        existed_counters_g += 1
                        y_true.add((slot_idx_map[ix], counter, counter))
                    else:
                        if corrections[slot_name]['fp'] > 0:
                            corrections[slot_name]['fp'] -= 1
                        else:
                            slot_sklearn_metric[slot_idx_map[ix]]['fp'] += 1
                    counter += 1
                if all_g_vals != ['-']:
                    remained_g_vals = len(all_g_vals) - existed_counters_g
                    for i in range(0, remained_g_vals):
                        if corrections[slot_name]['fn'] > 0:
                            corrections[slot_name]['fn'] -= 1
                        else:
                            slot_sklearn_metric[slot_idx_map[ix]]['fn'] += 1
                            y_true.add((slot_idx_map[ix], counter, counter))
                            counter += 1
            if p == '-' and refined_gold_slots[ix] != '-':
                all_g_vals = refined_gold_slots[ix].split(',')
                for g_v in all_g_vals:
                    if corrections[slot_name]['fn'] > 0:
                        corrections[slot_name]['fn'] -= 1
                    else:
                        slot_sklearn_metric[slot_idx_map[ix]]['fn'] += 1
                        y_true.add((slot_idx_map[ix], counter, counter))
                        counter += 1

    true_entities = y_true
    pred_entities = y_pred

    nb_correct = len(true_entities & pred_entities)
    nb_true = len(true_entities)
    nb_pred = len(pred_entities)

    recall_score = nb_correct / nb_true if nb_true > 0 else 0
    precision_score = nb_correct / nb_pred if nb_pred > 0 else 0

    p = nb_correct / nb_pred if nb_pred > 0 else 0
    r = nb_correct / nb_true if nb_true > 0 else 0
    f1_score = 2 * p * r / (p + r) if p + r > 0 else 0

    print(f'sklearn metric: precision = {precision_score}')
    print(f'sklearn metric: recall = {recall_score}')
    print(f'sklearn metric: F1 = {f1_score}')

    f1_avg, prec_avg, rec_agv = 0, 0, 0
    for slt in slot_sklearn_metric.keys():
        prec = slot_sklearn_metric[slt]['tp'] / (slot_sklearn_metric[slt]['tp'] + slot_sklearn_metric[slt]['fp'])
        rec = slot_sklearn_metric[slt]['tp'] / (slot_sklearn_metric[slt]['tp'] + slot_sklearn_metric[slt]['fn'])
        fone = 2 * (prec * rec) / (prec + rec)
        print(
            f"tp: {slot_sklearn_metric[slt]['tp']}  fp: {slot_sklearn_metric[slt]['fp']}   fn: {slot_sklearn_metric[slt]['fn']}")
        print(f'{slt}: precision = {prec}   recall = {rec}  f1 = {fone}')
        f1_avg += fone
        prec_avg += prec
        rec_agv += rec
    n = len(list(slot_sklearn_metric.keys()))
    print(f'avg F1 = {f1_avg / n}')
    print(f'avg precision = {prec_avg / n}')
    print(f'avg recall = {rec_agv / n}')
    print(f'topic_accuracy= {tp_top / len(predicted_data)}')


def build_prompt(data_pickle, inp_out_file, start_idx=-1, stop_idx=10000, sheet_name='gpt4o'):
    slot_map = {'تعداد خواهربرادر': 'Sibl', 'جنسیت': 'Gend', 'سن': 'Age', 'شغل': 'Job', 'سرگرمی': 'Hobby',
                'تعداد فرزندان': 'Child', 'محل سکونت': 'Resist', 'وضعیت تاهل': 'Marri', 'اسم': 'Name'}
    use_together = False
    if 'llama' in sheet_name:
        use_together = True

    with open(data_pickle, 'rb') as file:
        gold_data = pickle.load(file)

    workbook = load_workbook(inp_out_file)

    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        workbook.save(inp_out_file)
        predicted_data = []
    else:
        df_existing = pd.read_excel(inp_out_file, sheet_name=sheet_name)
        predicted_data = df_existing.values.tolist()

    for idx, data in enumerate(gold_data):
        if idx < start_idx:
            continue
        if idx >= stop_idx:
            break
        prev_utter = data[0]
        input_utter = data[1]
        print(predicted_data[idx])
        profile_predict, raw_response = get_processed_profile(input_utter, [prev_utter], perinfex, use_together,
                                                              predicted_data[idx][-1])

        gold_slot_values = {}
        gold_slots = data[2].split(',')
        for g_slot in gold_slots:
            g_key = g_slot.split('=')[0]
            g_val = g_slot.split('=')[1]
            if g_val != '-':
                g_vals = g_val.split('|')
                gold_slot_values[slot_map[g_key]] = g_vals

        profile_gold = process_slot_values(input_utter, gold_slot_values, perinfex)
        profile_gold['topic'] = data[3]
        profile_key_map = {'gender': 'جنسیت', 'marital status': 'تاهل', 'numberOfGirls': '#دختران',
                           'minNumberOfGirls': '#حداقل دختران',
                           'numberOfBoys': '#پسران', 'minNumberOfBoys': '#حداقل پسران', 'numberOfChild': '#فرزندان',
                           'minNumberOfChild': '#حداقل فرزندان',
                           'minNumberOfSisters': '#حداقل خواهران', 'numberOfSisters': '#خواهران',
                           'minNumberOfBrothers': '#حداقل برادران',
                           'numberOfBrothers': '#برادران', 'numberOfSibling': '#خواهربرادران',
                           'minNumberOfSibling': '#حداقل خواهربرادران',
                           'hobby': 'تفریح', 'job': 'شغل', 'age': 'سن', 'name': 'نام',
                           'city': 'شهر', 'country': 'کشور', 'region': 'جهت جغرافیا', 'province': 'استان',
                           'section': 'بخش'}

        predicted_slots = ''
        for key, val in profile_predict.items():
            if key == 'topic': continue
            if val == UNK or (isinstance(val, list) and len(val) == 0): val = '-'
            if isinstance(val, (str, int)):
                predicted_slots += profile_key_map[key] + ' = ' + str(val) + ' | '
            else:
                predicted_slots += profile_key_map[key] + ' = ' + ','.join(val) + ' | '
        predicted_slots = predicted_slots.strip('|')

        gold_slot_val = ''
        for key, val in profile_gold.items():
            if key == 'topic': continue
            if val == UNK or (isinstance(val, list) and len(val) == 0): val = '-'
            if isinstance(val, (str, int)):
                gold_slot_val += profile_key_map[key] + ' = ' + str(val) + ' | '
            else:
                gold_slot_val += profile_key_map[key] + ' = ' + ','.join(val) + ' | '
        gold_slot_val = gold_slot_val.strip('|')

        print(input_utter)
        print('prediction: ', predicted_slots)
        print('pred topic: ', profile_predict['topic'])
        print('-' * 10)
        predicted_data.append(
            (data[0], data[1], gold_slot_val, profile_gold['topic'], predicted_slots, profile_predict['topic'],
             raw_response))
    new_name = inp_out_file.split('.')[0] + '_modified' + '.xlsx'
    writer = pd.ExcelWriter(new_name, engine='openpyxl', mode='w')
    df = pd.DataFrame(predicted_data, columns=['prev_utter', 'utterance', 'gold_slots', 'gold_topic', 'pred_slots',
                                               'pred_topic', 'raw_response'])  # Convert your data to a DataFrame
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook.save(inp_out_file)
    writer.close()


def get_mine_nlu_result(data_pickle, inp_out_file, start_idx=-1, stop_idx=10000, sheet_name='mine_nlu'):
    slot_map = {'تعداد خواهربرادر': 'Sibl', 'جنسیت': 'Gend', 'سن': 'Age', 'شغل': 'Job', 'سرگرمی': 'Hobby',
                'تعداد فرزندان': 'Child', 'محل سکونت': 'Resist', 'وضعیت تاهل': 'Marri', 'اسم': 'Name'}

    with open(data_pickle, 'rb') as file:
        gold_data = pickle.load(file)

    workbook = load_workbook(inp_out_file)

    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        workbook.save(inp_out_file)
        predicted_data = []
    else:
        df_existing = pd.read_excel(inp_out_file, sheet_name=sheet_name)
        predicted_data = df_existing.values.tolist()

    for idx, data in enumerate(gold_data):
        if idx < start_idx:
            continue
        if idx >= stop_idx:
            break
        prev_utter = data[0]
        input_utter = data[1]

        gold_slot_values = {}
        gold_slots = data[2].split(',')
        gold_topic = data[3]
        for g_slot in gold_slots:
            g_key = g_slot.split('=')[0]
            g_val = g_slot.split('=')[1]
            if g_val != '-':
                g_vals = g_val.split('|')
                gold_slot_values[slot_map[g_key]] = g_vals

        nlu_output = get_nlu_prediction([(input_utter, prev_utter)])[0]
        slot_vals, subj = get_slot_values(nlu_output)

        predicted_slots = ''
        for per_slot, eng_slot in slot_map.items():
            if eng_slot not in slot_vals.keys():
                val = '-'
            else:
                val = slot_vals[eng_slot]
            if isinstance(val, (str, int)):
                predicted_slots += per_slot + ' = ' + str(val) + ' | '
            else:
                predicted_slots += per_slot + ' = ' + ','.join(val) + ' | '

        gold_slot_val = ''
        for per_slot, eng_slot in slot_map.items():
            if eng_slot not in gold_slot_values.keys():
                val = '-'
            else:
                val = gold_slot_values[eng_slot]
            if isinstance(val, (str, int)):
                gold_slot_val += per_slot + ' = ' + str(val) + ' | '
            else:
                gold_slot_val += per_slot + ' = ' + ','.join(val) + ' | '

        predicted_data.append(
            (data[0], data[1], gold_slot_val, gold_topic, predicted_slots, subj))

    writer = pd.ExcelWriter(inp_out_file, engine='openpyxl', mode='a', if_sheet_exists='replace')
    df = pd.DataFrame(predicted_data, columns=['prev_utter', 'utterance', 'gold_slots', 'gold_topic', 'pred_slots',
                                               'pred_topic'])  # Convert your data to a DataFrame
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook.save(inp_out_file)
    writer.close()


if __name__ == '__main__':
    # get_mine_nlu_result('test_data.pkl', 'Output_excel_file.xlsx', start_idx=-1, stop_idx=730, sheet_name='mine_nlu')
    get_gold_preds('Input_excel_file.xlsx', 'mine_nlu', check_single_topic=True)  # 'mine_nlu'
    # get_gold_preds_llm('Input_excel_file.xlsx', 'gpt4o', check_single_topic=True)  # 'gpt4o/mine_nlu'/'phi-4'/'llama-8B-improve'